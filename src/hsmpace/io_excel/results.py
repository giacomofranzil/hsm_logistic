"""Writing the results to a separate workbook.

The input file always stays read only: results are never written inside the
file the user fills in, to avoid workbooks floating around with new inputs and
stale results.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..core.analysis import GapAnalysis, mass_balance
from ..core.model import Case, MassFlowDeviation
from ..core.simulate import PieceResult
from ..core.studies import MonteCarloResult, PacingPoint

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _table(wb: Workbook, title: str, header: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title)
    for col, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(34, len(name) + 6))
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    ws.freeze_panes = "A2"


def write_results(
    path: str | Path,
    case: Case,
    results: list[PieceResult],
    analyses: list[GapAnalysis],
    deviations: tuple[MassFlowDeviation, ...] = (),
    curve: list[PacingPoint] | None = None,
    min_pacing: PacingPoint | None = None,
    mc: MonteCarloResult | None = None,
) -> Path:
    path = Path(path)
    wb = Workbook()
    wb.remove(wb.active)

    summary: list[list] = [
        ["Mill", case.info.get("mill_name", "")],
        ["Nominal pacing [s]", case.settings.pacing],
        ["Minimum gap required [m]", case.settings.gap_min],
        ["Pieces simulated", len(results)],
        ["Product sequence", ", ".join(case.piece_products)],
        ["Coiler sequence", ", ".join(cid or "-" for cid in case.piece_coiler_ids)],
    ]
    if analyses:
        worst = min(analyses, key=lambda a: a.min_gap)
        summary += [
            ["Minimum gap of the sequence [m]", round(worst.min_gap, 2)],
            ["Critical pair", f"{worst.front_id} / {worst.rear_id}"],
            ["Critical position [m]", round(worst.critical.x, 1) if worst.critical else ""],
            ["Critical instant [s]", round(worst.critical.t, 1) if worst.critical else ""],
            ["Critical section", worst.critical.section if worst.critical else ""],
            ["Time gap at the critical point [s]", round(worst.headway, 1) if worst.headway else ""],
            ["Outcome", "feasible" if worst.ok else "VIOLATION"],
        ]
    else:
        summary.append(["Outcome", "the pieces never interact"])
    if min_pacing is not None:
        summary += [
            ["Minimum feasible pacing [s]", round(min_pacing.pacing, 1)],
            ["Binding constraint at the minimum pacing", min_pacing.section or min_pacing.equipment],
        ]
    if mc is not None:
        summary += [
            ["Monte Carlo: runs", mc.runs],
            ["Monte Carlo: violations", mc.violations],
            ["Monte Carlo: violation rate [%]", round(100 * mc.violation_rate, 2)],
            ["Monte Carlo: mean gap [m]", round(mc.mean, 2)],
            ["Monte Carlo: 5th percentile gap [m]", round(mc.percentile(0.05), 2)],
        ]
    _table(wb, "Summary", ["Item", "Value"], summary)

    _table(
        wb,
        "Events",
        ["piece", "t [s]", "event", "equipment", "x [m]", "detail"],
        [
            [r.piece_id, round(e.t, 3), e.kind, e.equipment_id, round(e.x, 2), e.detail]
            for r in results
            for e in r.events
        ],
    )

    _table(
        wb,
        "Occupancy",
        ["piece", "equipment", "pass", "in [s]", "out [s]", "duration [s]"],
        [
            [r.piece_id, o.equipment_id, o.pass_no, round(o.t_in, 2), round(o.t_out, 2), round(o.duration, 2)]
            for r in results
            for o in r.occupancy
        ],
    )

    _table(
        wb,
        "Gap",
        [
            "front piece",
            "rear piece",
            "minimum gap [m]",
            "t [s]",
            "x [m]",
            "section",
            "equipment",
            "time gap [s]",
            "first violation [s]",
            "outcome",
        ],
        [
            [
                a.front_id,
                a.rear_id,
                round(a.min_gap, 2),
                round(a.critical.t, 2) if a.critical else "",
                round(a.critical.x, 2) if a.critical else "",
                a.critical.section if a.critical else "",
                a.critical.equipment if a.critical else "",
                round(a.headway, 2) if a.headway is not None else "",
                round(a.t_first_violation, 2) if a.t_first_violation is not None else "",
                "ok" if a.ok else "VIOLATION",
            ]
            for a in analyses
        ],
    )

    if curve:
        _table(
            wb,
            "PacingCurve",
            ["pacing [s]", "minimum gap [m]", "feasible", "critical t [s]", "critical x [m]", "section", "pair"],
            [
                [
                    round(p.pacing, 2),
                    round(p.min_gap, 2) if p.min_gap != float("inf") else "no interaction",
                    "yes" if p.feasible else "no",
                    round(p.t_critical, 2) if p.t_critical is not None else "",
                    round(p.x_critical, 2) if p.x_critical is not None else "",
                    p.section,
                    p.pair,
                ]
                for p in curve
            ],
        )

    _table(
        wb,
        "MassBalance",
        ["product", "kinematic length [m]", "geometric length [m]", "deviation [%]", "outcome"],
        [
            [c.piece_id, round(c.length_kinematic, 2), round(c.length_geometric, 2), round(c.error_pct, 4), "ok" if c.ok else "WARNING"]
            for c in mass_balance(results)
        ],
    )

    if deviations:
        _table(
            wb,
            "MassFlowTandem",
            ["product", "pass", "stand", "v entered [m/s]", "v from balance [m/s]", "deviation [%]"],
            [
                [d.product_id, d.pass_no, d.equipment_id, round(d.v_input, 4), round(d.v_massflow, 4), round(d.deviation_pct, 3)]
                for d in deviations
            ],
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
