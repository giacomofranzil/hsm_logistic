"""Lettura del workbook di input con validazione severa.

Ogni problema viene riportato con foglio, cella e motivo: un input ambiguo
viene rifiutato con un messaggio comprensibile, mai con un traceback.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..core.model import (
    FWD,
    REV,
    Case,
    Equipment,
    Line,
    Product,
    RollingPass,
    Section,
    SimSettings,
    SpeedEvent,
    validate_case,
)
from . import schema as S

_TRUE = {"si", "s", "true", "vero", "x", "yes", "y", "1"}
_FALSE = {"no", "n", "false", "falso", "0", ""}


@dataclass(frozen=True)
class ValidationIssue:
    sheet: str
    cell: str
    message: str

    def __str__(self) -> str:
        where = f"{self.sheet}!{self.cell}" if self.cell else self.sheet
        return f"{where}: {self.message}"


class ValidationError(Exception):
    def __init__(self, issues: list[ValidationIssue]) -> None:
        self.issues = issues
        head = "\n".join(f"  - {i}" for i in issues[:40])
        extra = "" if len(issues) <= 40 else f"\n  ... e altri {len(issues) - 40} problemi"
        super().__init__(f"Input non valido ({len(issues)} problemi):\n{head}{extra}")


class _Collector:
    def __init__(self) -> None:
        self.issues: list[ValidationIssue] = []

    def add(self, sheet: str, cell: str, message: str) -> None:
        self.issues.append(ValidationIssue(sheet, cell, message))

    def raise_if_any(self) -> None:
        if self.issues:
            raise ValidationError(self.issues)


class _Table:
    """Foglio tabellare con intestazione in riga 1."""

    def __init__(self, ws: Any, columns: list[str], issues: _Collector) -> None:
        self.ws = ws
        self.name = ws.title
        self.issues = issues
        header = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is not None and str(value).strip():
                header[str(value).strip()] = col
        self.header = header
        for name in columns:
            if name not in header:
                issues.add(self.name, "1", f"colonna mancante: {name!r}")

    def rows(self) -> list[int]:
        out = []
        for row in range(2, self.ws.max_row + 1):
            if any(
                self.ws.cell(row=row, column=col).value not in (None, "")
                for col in self.header.values()
            ):
                out.append(row)
        return out

    def ref(self, row: int, name: str) -> str:
        col = self.header.get(name)
        return f"{get_column_letter(col)}{row}" if col else f"A{row}"

    def raw(self, row: int, name: str) -> Any:
        col = self.header.get(name)
        if col is None:
            return None
        value = self.ws.cell(row=row, column=col).value
        return value.strip() if isinstance(value, str) else value

    def text(self, row: int, name: str, required: bool = True, default: str = "") -> str:
        value = self.raw(row, name)
        if value is None or value == "":
            if required:
                self.issues.add(self.name, self.ref(row, name), f"{name}: valore obbligatorio")
            return default
        return str(value).strip()

    def number(
        self,
        row: int,
        name: str,
        required: bool = True,
        default: float | None = None,
        minimum: float | None = None,
    ) -> float | None:
        value = self.raw(row, name)
        if value is None or value == "":
            if required:
                self.issues.add(self.name, self.ref(row, name), f"{name}: valore obbligatorio")
            return default
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            try:
                value = float(str(value).replace(",", "."))
            except ValueError:
                self.issues.add(
                    self.name,
                    self.ref(row, name),
                    f"{name}: {value!r} non e' un numero",
                )
                return default
        value = float(value)
        if minimum is not None and value < minimum:
            self.issues.add(
                self.name,
                self.ref(row, name),
                f"{name}: {value:g} sotto il minimo ammesso ({minimum:g})",
            )
        return value

    def integer(self, row: int, name: str, required: bool = True, default: int = 0) -> int:
        value = self.number(row, name, required, float(default))
        return int(round(value)) if value is not None else default

    def flag(self, row: int, name: str, default: bool = False) -> bool:
        value = self.raw(row, name)
        if value is None or value == "":
            return default
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in _TRUE:
            return True
        if text in _FALSE:
            return False
        self.issues.add(
            self.name, self.ref(row, name), f"{name}: {value!r} non e' SI/NO"
        )
        return default

    def direction(self, row: int, name: str, default: int = FWD) -> int:
        value = self.raw(row, name)
        if value is None or value == "":
            return default
        text = str(value).strip().lower()
        if text in {"fwd", "avanti", "+1", "1", "diretta"}:
            return FWD
        if text in {"rev", "indietro", "-1", "inversa"}:
            return REV
        self.issues.add(
            self.name, self.ref(row, name), f"{name}: {value!r} non e' fwd/rev"
        )
        return default


def _key_values(ws: Any) -> dict[str, tuple[Any, str]]:
    out: dict[str, tuple[Any, str]] = {}
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        if key is None or not str(key).strip():
            continue
        value = ws.cell(row=row, column=2).value
        out[str(key).strip()] = (value, f"B{row}")
    return out


def read_case(path: str | Path) -> Case:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"file non trovato: {path}")

    issues = _Collector()
    wb = load_workbook(path, data_only=True, read_only=False)

    for name in (S.SHEET_INFO, S.SHEET_LAYOUT, S.SHEET_SECTIONS, S.SHEET_PRODUCTS, S.SHEET_PASSES):
        if name not in wb.sheetnames:
            issues.add(name, "", "foglio mancante nel workbook")
    issues.raise_if_any()

    info_raw = _key_values(wb[S.SHEET_INFO])
    info = {k: ("" if v is None else str(v)) for k, (v, _) in info_raw.items()}
    version = info.get("schema_version", "").strip()
    if version and version != S.SCHEMA_VERSION:
        issues.add(
            S.SHEET_INFO,
            info_raw["schema_version"][1],
            f"schema_version {version!r} non supportata, attesa {S.SCHEMA_VERSION!r}",
        )

    locations: dict[str, tuple[str, str]] = {}
    equipment = _read_layout(wb[S.SHEET_LAYOUT], issues, locations)
    sections = _read_sections(wb[S.SHEET_SECTIONS], equipment, issues, locations)
    products = _read_products(wb[S.SHEET_PRODUCTS], wb[S.SHEET_PASSES], issues, locations)
    if S.SHEET_SIM in wb.sheetnames:
        settings = _read_settings(wb[S.SHEET_SIM], issues, locations)
    else:
        settings = SimSettings()
    issues.raise_if_any()

    case = Case(
        line=Line(tuple(equipment), tuple(sections)),
        products=tuple(products),
        settings=settings,
        info=info,
    )

    for problem in validate_case(case):
        sheet, cell = locations.get(problem.locator, ("modello", ""))
        issues.add(sheet, cell, problem.message)
    issues.raise_if_any()
    return case


def _read_layout(
    ws: Any, issues: _Collector, locations: dict[str, tuple[str, str]]
) -> list[Equipment]:
    table = _Table(ws, [c[0] for c in S.LAYOUT_COLUMNS], issues)
    out: list[Equipment] = []
    for row in table.rows():
        kind = table.text(row, "kind").lower()
        locations[f"equipment:{table.text(row, 'equipment_id', required=False)}"] = (
            table.name,
            table.ref(row, "equipment_id"),
        )
        if kind not in ("start", "stand", "coiler", "marker"):
            issues.add(
                table.name,
                table.ref(row, "kind"),
                f"kind {kind!r} non valido: usare start, stand, coiler o marker",
            )
        out.append(
            Equipment(
                id=table.text(row, "equipment_id"),
                kind=kind,
                x=table.number(row, "x_m") or 0.0,
                accel=table.number(row, "accel_mps2", required=False, default=1.0, minimum=1e-6)
                or 1.0,
                group=table.text(row, "group", required=False),
                label=table.text(row, "label", required=False),
            )
        )
    if not out:
        issues.add(table.name, "A2", "nessuna apparecchiatura definita")
    return out


def _read_sections(
    ws: Any,
    equipment: list[Equipment],
    issues: _Collector,
    locations: dict[str, tuple[str, str]],
) -> list[Section]:
    columns = [c[0] for c in S.SECTION_COLUMNS]
    table = _Table(ws, columns, issues)
    by_id = {e.id: e for e in equipment}
    out: list[Section] = []
    x_prev_end = 0.0

    for row in table.rows():
        section_id = table.text(row, "section_id")
        locations[f"section:{section_id}"] = (table.name, table.ref(row, "section_id"))
        ref = table.text(row, "start_ref", required=False)
        offset = table.number(row, "start_offset_m", required=False, default=0.0) or 0.0
        if not ref:
            x_start = offset
        elif ref.lower() == "prev":
            x_start = x_prev_end + offset
        elif ref in by_id:
            x_start = by_id[ref].x + offset
        else:
            issues.add(
                table.name,
                table.ref(row, "start_ref"),
                f"riferimento {ref!r} sconosciuto: usare un equipment_id, 'prev' o lasciare vuoto",
            )
            x_start = offset

        length = table.number(row, "length_m", minimum=1e-6) or 0.0
        direction = table.direction(row, "direction")
        during_pass = table.flag(row, "during_pass", default=False)

        events: list[SpeedEvent] = []
        last_d: float | None = None
        for i in range(1, S.MAX_EVENTS_PER_SECTION + 1):
            d_raw = table.raw(row, f"d{i}_m")
            v_raw = table.raw(row, f"v{i}_mps")
            if d_raw in (None, "") and v_raw in (None, ""):
                continue
            if d_raw in (None, "") or v_raw in (None, ""):
                issues.add(
                    table.name,
                    table.ref(row, f"d{i}_m"),
                    f"evento {i}: distanza e velocita' vanno indicate entrambe",
                )
                continue
            distance = table.number(row, f"d{i}_m", minimum=0.0) or 0.0
            speed = table.number(row, f"v{i}_mps", minimum=0.0) or 0.0
            accel = table.number(row, f"a{i}_mps2", required=False, default=None, minimum=1e-6)
            if distance > length + 1e-9:
                issues.add(
                    table.name,
                    table.ref(row, f"d{i}_m"),
                    f"evento {i}: distanza {distance:g} m oltre la lunghezza della sezione "
                    f"({length:g} m)",
                )
            if last_d is not None and distance < last_d - 1e-9:
                issues.add(
                    table.name,
                    table.ref(row, f"d{i}_m"),
                    f"evento {i}: distanze non crescenti ({distance:g} dopo {last_d:g})",
                )
            last_d = distance
            events.append(
                SpeedEvent(
                    id=f"{section_id}-{i}",
                    section_id=section_id,
                    x_trigger=x_start + distance,
                    v_target=speed,
                    accel=accel,
                    direction=direction,
                    during_pass=during_pass,
                )
            )

        out.append(
            Section(
                id=section_id,
                x_start=x_start,
                length=length,
                label=table.text(row, "label", required=False),
                events=tuple(events),
            )
        )
        x_prev_end = x_start + length

    return out


def _read_products(
    ws_products: Any,
    ws_passes: Any,
    issues: _Collector,
    locations: dict[str, tuple[str, str]],
) -> list[Product]:
    table = _Table(ws_products, [c[0] for c in S.PRODUCT_COLUMNS], issues)
    passes_table = _Table(ws_passes, [c[0] for c in S.PASS_COLUMNS], issues)

    by_product: dict[str, list[RollingPass]] = {}
    for row in passes_table.rows():
        product_id = passes_table.text(row, "product_id")
        pass_no = passes_table.integer(row, "pass_no")
        locations[f"pass:{product_id}:{pass_no}"] = (
            passes_table.name,
            passes_table.ref(row, "pass_no"),
        )
        rp = RollingPass(
            product_id=product_id,
            pass_no=pass_no,
            equipment_id=passes_table.text(row, "equipment_id"),
            direction=passes_table.direction(row, "direction"),
            h_in=passes_table.number(row, "h_in_mm", minimum=1e-6) or 0.0,
            h_out=passes_table.number(row, "h_out_mm", minimum=1e-6) or 0.0,
            w_in=passes_table.number(row, "w_in_mm", minimum=1e-6) or 0.0,
            w_out=passes_table.number(row, "w_out_mm", minimum=1e-6) or 0.0,
            v_exit=passes_table.number(row, "v_exit_mps", minimum=1e-6) or 0.0,
            reversing_delay=passes_table.number(
                row, "reversing_delay_s", required=False, default=0.0, minimum=0.0
            )
            or 0.0,
            approach_v=passes_table.number(
                row, "approach_v_mps", required=False, default=None, minimum=1e-6
            ),
            master=passes_table.flag(row, "master"),
            zoom_pct=passes_table.number(row, "zoom_pct", required=False, default=0.0) or 0.0,
            zoom_trigger=passes_table.number(row, "zoom_trigger_m", required=False, default=0.0)
            or 0.0,
            zoom_accel=passes_table.number(
                row, "zoom_accel_mps2", required=False, default=None, minimum=1e-6
            ),
        )
        by_product.setdefault(product_id, []).append(rp)

    out: list[Product] = []
    for row in table.rows():
        product_id = table.text(row, "product_id")
        locations[f"product:{product_id}"] = (table.name, table.ref(row, "product_id"))
        passes = sorted(by_product.pop(product_id, []), key=lambda p: p.pass_no)
        if not passes:
            issues.add(
                table.name,
                table.ref(row, "product_id"),
                f"nessuna passata nel foglio {S.SHEET_PASSES} per il prodotto {product_id!r}",
            )
        numbers = [p.pass_no for p in passes]
        for dup in sorted({n for n in numbers if numbers.count(n) > 1}):
            issues.add(
                passes_table.name, "", f"prodotto {product_id}: numero passata duplicato {dup}"
            )
        out.append(
            Product(
                id=product_id,
                slab_thk=table.number(row, "slab_thk_mm", minimum=1e-6) or 0.0,
                slab_wid=table.number(row, "slab_wid_mm", minimum=1e-6) or 0.0,
                slab_len=table.number(row, "slab_len_m", minimum=1e-6) or 0.0,
                label=table.text(row, "label", required=False),
                grade=table.text(row, "grade", required=False),
                passes=tuple(passes),
            )
        )

    for orphan in by_product:
        issues.add(
            passes_table.name,
            "",
            f"passate riferite al prodotto {orphan!r}, assente dal foglio {S.SHEET_PRODUCTS}",
        )
    return out


def _read_settings(
    ws: Any, issues: _Collector, locations: dict[str, tuple[str, str]]
) -> SimSettings:
    raw = _key_values(ws)
    defaults = SimSettings()
    for key, (_value, cell) in raw.items():
        locations[f"setting:{key}"] = (ws.title, cell)

    def num(key: str, fallback: float) -> float:
        if key not in raw:
            return fallback
        value, cell = raw[key]
        if value is None or value == "":
            return fallback
        try:
            return float(str(value).replace(",", "."))
        except ValueError:
            issues.add(ws.title, cell, f"{key}: {value!r} non e' un numero")
            return fallback

    def flag(key: str, fallback: bool) -> bool:
        if key not in raw:
            return fallback
        value, cell = raw[key]
        if value is None or value == "":
            return fallback
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if text in _TRUE:
            return True
        if text in _FALSE:
            return False
        issues.add(ws.title, cell, f"{key}: {value!r} non e' SI/NO")
        return fallback

    sequence: tuple[str, ...] = ()
    if "piece_products" in raw:
        value, _ = raw["piece_products"]
        if value:
            sequence = tuple(p.strip() for p in str(value).split(",") if p.strip())

    return SimSettings(
        pacing=num("pacing_s", defaults.pacing),
        n_pieces=int(num("n_pieces", defaults.n_pieces)),
        piece_products=sequence,
        gap_min=num("gap_min_m", defaults.gap_min),
        pacing_scan_min=num("pacing_scan_min_s", defaults.pacing_scan_min),
        pacing_scan_max=num("pacing_scan_max_s", defaults.pacing_scan_max),
        pacing_scan_steps=int(num("pacing_scan_steps", defaults.pacing_scan_steps)),
        mc_runs=int(num("mc_runs", defaults.mc_runs)),
        mc_speed_tol_pct=num("mc_speed_tol_pct", defaults.mc_speed_tol_pct),
        mc_delay_sigma=num("mc_delay_sigma_s", defaults.mc_delay_sigma),
        mc_release_sigma=num("mc_release_sigma_s", defaults.mc_release_sigma),
        mc_seed=int(num("mc_seed", defaults.mc_seed)),
        max_time=num("max_time_s", defaults.max_time),
        time_axis_down=flag("time_axis_down", defaults.time_axis_down),
    )
