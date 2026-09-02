"""Writing a Case to an Excel workbook: empty template or example case."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..core.model import FWD, Case, Section
from . import schema as S

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=12)


def _write_header(ws: Worksheet, columns: list[tuple[str, str]]) -> None:
    for col, (name, note) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if note:
            cell.comment = Comment(note, "hsmpace")
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(26, len(name) + 8))
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def _section_columns() -> list[tuple[str, str]]:
    columns = list(S.SECTION_COLUMNS)
    for i in range(1, S.MAX_EVENTS_PER_SECTION + 1):
        for name, note in S.SECTION_EVENT_COLUMNS:
            columns.append((name.format(i), note.format(i)))
    return columns


def _bool(value: bool) -> str:
    return "YES" if value else "NO"


def _direction(value: int) -> str:
    return "fwd" if value == FWD else "rev"


def _section_start_ref(case: Case, section: Section) -> tuple[str, float]:
    """Readable reference for the section start: the closest equipment upstream."""
    upstream = [e for e in case.line.equipment if e.x <= section.x_start + 1e-9]
    if not upstream:
        return "", section.x_start
    anchor = max(upstream, key=lambda e: e.x)
    return anchor.id, round(section.x_start - anchor.x, 6)


def write_case(case: Case, path: str | Path, include_data: bool = True) -> Path:
    """Write the workbook. With ``include_data`` false it produces the empty template."""
    path = Path(path)
    wb = Workbook()

    ws = wb.active
    ws.title = S.SHEET_GUIDE
    ws.column_dimensions["A"].width = 110
    for row, (text, is_title) in enumerate(S.GUIDE_TEXT, start=1):
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if is_title:
            cell.font = _TITLE_FONT
        if len(text) > 90:
            ws.row_dimensions[row].height = 15 * (len(text) // 90 + 1)

    ws = wb.create_sheet(S.SHEET_INFO)
    _write_header(ws, [("key", "Key"), ("value", "Value")])
    info = {"schema_version": S.SCHEMA_VERSION, "mill_name": "", "notes": ""}
    if include_data:
        info.update(case.info)
    info["schema_version"] = S.SCHEMA_VERSION
    for row, (key, value) in enumerate(info.items(), start=2):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)

    ws = wb.create_sheet(S.SHEET_LAYOUT)
    _write_header(ws, S.LAYOUT_COLUMNS)
    if include_data:
        for row, eq in enumerate(case.line.equipment, start=2):
            # acceleration is only read on stand and coiler rows
            accel = eq.accel if eq.kind in ("stand", "coiler") else None
            for col, value in enumerate(
                [eq.id, eq.kind, eq.x, accel, eq.group, eq.label], start=1
            ):
                ws.cell(row=row, column=col, value=value)

    ws = wb.create_sheet(S.SHEET_SECTIONS)
    _write_header(ws, _section_columns())
    if include_data:
        for row, sec in enumerate(case.line.sections, start=2):
            ref, offset = _section_start_ref(case, sec)
            during = any(e.during_pass for e in sec.events)
            direction = sec.events[0].direction if sec.events else FWD
            values = [
                sec.id,
                sec.label,
                ref,
                offset,
                sec.length,
                _direction(direction),
                _bool(during),
                sec.accel,
            ]
            for i, ev in enumerate(sec.events[: S.MAX_EVENTS_PER_SECTION]):
                values.extend(
                    [round(ev.x_trigger - sec.x_start, 6), ev.v_target, ev.accel]
                )
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)

    ws = wb.create_sheet(S.SHEET_PRODUCTS)
    _write_header(ws, S.PRODUCT_COLUMNS)
    if include_data:
        for row, p in enumerate(case.products, start=2):
            for col, value in enumerate(
                [p.id, p.label, p.grade, p.slab_thk, p.slab_wid, p.slab_len], start=1
            ):
                ws.cell(row=row, column=col, value=value)

    ws = wb.create_sheet(S.SHEET_PASSES)
    _write_header(ws, S.PASS_COLUMNS)
    if include_data:
        row = 2
        for product in case.products:
            for rp in product.passes:
                values = [
                    product.id,
                    rp.pass_no,
                    rp.equipment_id,
                    _direction(rp.direction),
                    rp.h_in,
                    rp.h_out,
                    rp.w_in,
                    rp.w_out,
                    rp.v_exit_input if rp.v_exit_input is not None else rp.v_exit,
                    rp.reversing_delay or None,
                    rp.reversing_clearance or None,
                    rp.approach_v,
                    _bool(rp.master) if rp.master else "",
                    rp.zoom_pct or None,
                    rp.zoom_trigger or None,
                    rp.zoom_accel,
                ]
                for col, value in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=value)
                row += 1

    ws = wb.create_sheet(S.SHEET_SIM)
    _write_header(ws, [("key", "Key"), ("value", "Value"), ("note", "Description")])
    ws.column_dimensions["C"].width = 70
    s = case.settings
    values = {
        "pacing_s": s.pacing,
        "n_pieces": s.n_pieces,
        "piece_products": ",".join(s.piece_products),
        "gap_min_m": s.gap_min,
        "pacing_scan_min_s": s.pacing_scan_min,
        "pacing_scan_max_s": s.pacing_scan_max,
        "pacing_scan_steps": s.pacing_scan_steps,
        "mc_runs": s.mc_runs,
        "mc_speed_tol_pct": s.mc_speed_tol_pct,
        "mc_delay_sigma_s": s.mc_delay_sigma,
        "mc_release_sigma_s": s.mc_release_sigma,
        "mc_seed": s.mc_seed,
        "table_accel_mps2": s.table_accel,
        "coiler_v_final_mps": s.coiler_v_final,
        "max_time_s": s.max_time,
        "time_axis_down": _bool(s.time_axis_down),
    }
    for row, (key, default, note) in enumerate(S.SIM_KEYS, start=2):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=values[key] if include_data else default)
        cell = ws.cell(row=row, column=3, value=note)
        cell.alignment = Alignment(wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
