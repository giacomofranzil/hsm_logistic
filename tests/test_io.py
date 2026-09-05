"""Checks of the Excel layer, of the JSON contract and of the tracking import."""

from __future__ import annotations

import json

import pytest
from openpyxl import load_workbook

from hsmpace.core.analysis import analyse_sequence
from hsmpace.core.contract import case_from_dict, case_to_dict, report_to_dict
from hsmpace.core.model import harmonise_tandem_speeds
from hsmpace.core.studies import base_results, gap_vs_pacing, min_feasible_pacing, sequence
from hsmpace.core.tracking import parse_tracking
from hsmpace.example import example_case
from hsmpace.io_excel import ValidationError, read_case, write_case, write_results


def test_excel_round_trip_preserves_the_case(tmp_path):
    original = example_case()
    path = write_case(original, tmp_path / "case.xlsx")
    loaded = read_case(path)

    assert loaded.line.equipment == original.line.equipment
    assert loaded.line.sections == original.line.sections
    assert loaded.products == original.products
    assert loaded.settings == original.settings


def test_the_empty_template_has_the_sheets_and_headers(tmp_path):
    path = write_case(example_case(), tmp_path / "template.xlsx", include_data=False)
    wb = load_workbook(path)
    assert {"Guide", "Info", "Layout", "Sections", "Products", "PassSchedule", "Simulation"} <= set(
        wb.sheetnames
    )
    assert wb["Layout"]["A1"].value == "equipment_id"
    header = [c.value for c in wb["Sections"][1]]
    assert "d1_m" in header and "accel_mps2" in header
    assert wb["Layout"].max_row == 1, "the template must contain no data"
    guide = "\n".join(str(row[0].value or "") for row in wb["Guide"].iter_rows(min_col=1, max_col=1))
    assert "tandem slows down anyway" in guide
    assert "as late as possible" in guide
    assert "offline model TRoll" in guide
    assert "whichever coiler takes the strip" in guide
    assert "coiler_pattern" in guide
    assert "Several coilers" in guide


def test_json_round_trip_preserves_the_case():
    original = example_case()
    payload = json.loads(json.dumps(case_to_dict(original)))
    loaded = case_from_dict(payload)

    assert loaded.line.equipment == original.line.equipment
    assert loaded.products == original.products
    assert loaded.settings == original.settings


def test_parsing_errors_point_at_sheet_and_cell(tmp_path):
    path = write_case(example_case(), tmp_path / "broken.xlsx")
    wb = load_workbook(path)
    wb["Layout"]["C5"] = "twenty five"
    wb.save(path)

    with pytest.raises(ValidationError) as exc:
        read_case(path)
    issue = exc.value.issues[0]
    assert issue.sheet == "Layout"
    assert issue.cell == "C5"
    assert "is not a number" in issue.message


def test_model_errors_trace_back_to_the_pass_row(tmp_path):
    path = write_case(example_case(), tmp_path / "broken2.xlsx")
    wb = load_workbook(path)
    wb["PassSchedule"]["F4"] = 200.0  # h_out larger than h_in
    wb.save(path)

    with pytest.raises(ValidationError) as exc:
        read_case(path)
    reduction = [i for i in exc.value.issues if "invalid reduction" in i.message]
    assert reduction and reduction[0].sheet == "PassSchedule"
    assert reduction[0].cell == "B4"


def test_an_unsupported_schema_version_is_rejected(tmp_path):
    path = write_case(example_case(), tmp_path / "old.xlsx")
    wb = load_workbook(path)
    wb["Info"]["B2"] = "0"
    wb.save(path)

    with pytest.raises(ValidationError, match="schema_version"):
        read_case(path)


def test_a_section_event_beyond_the_length_is_rejected(tmp_path):
    path = write_case(example_case(), tmp_path / "section.xlsx")
    wb = load_workbook(path)
    ws = wb["Sections"]
    col = [c.value for c in ws[1]].index("d1_m") + 1
    ws.cell(row=4, column=col, value=999.0)
    wb.save(path)

    with pytest.raises(ValidationError, match="beyond the section length"):
        read_case(path)


def test_an_older_workbook_without_optional_columns_still_loads(tmp_path):
    """Columns added after the first schema version must not break existing files."""
    path = write_case(example_case(), tmp_path / "older.xlsx")
    wb = load_workbook(path)
    ws = wb["PassSchedule"]
    header = [c.value for c in ws[1]]
    ws.delete_cols(header.index("reversing_clearance_m") + 1)
    wb.save(path)

    loaded = read_case(path)
    assert all(p.reversing_clearance == 0.0 for p in loaded.products[0].passes)


def test_an_older_workbook_without_coiler_pattern_still_loads(tmp_path):
    """A single-coiler file written before the pattern key remains valid."""
    from dataclasses import replace

    from hsmpace.core.model import Line

    original = example_case()
    equipment = tuple(e for e in original.line.equipment if e.id != "DC2")
    one = replace(
        original,
        line=Line(equipment, original.line.sections),
        settings=replace(original.settings, coiler_pattern=()),
    )
    path = write_case(one, tmp_path / "no_pattern.xlsx")
    wb = load_workbook(path)
    ws = wb["Simulation"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "coiler_pattern":
            ws.delete_rows(row)
            break
    wb.save(path)

    loaded = read_case(path)
    assert loaded.settings.coiler_pattern == ()
    assert loaded.piece_coiler_ids == ("DC1",) * loaded.settings.n_pieces


def test_two_coilers_in_excel_require_the_pattern(tmp_path):
    path = write_case(example_case(), tmp_path / "no_cycle.xlsx")
    wb = load_workbook(path)
    ws = wb["Simulation"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "coiler_pattern":
            ws.cell(row=row, column=2).value = ""
            break
    wb.save(path)

    with pytest.raises(ValidationError, match="coiler_pattern"):
        read_case(path)


def test_a_misplaced_reversal_value_is_only_a_warning(tmp_path):
    """The delay belongs to the pass the reversal comes after, not to any pass."""
    path = write_case(example_case(), tmp_path / "misplaced.xlsx")
    wb = load_workbook(path)
    ws = wb["PassSchedule"]
    col = [c.value for c in ws[1]].index("reversing_delay_s") + 1
    ws.cell(row=14, column=col, value=9.0)  # last pass, no reversal follows it
    wb.save(path)

    loaded = read_case(path)
    assert any("no reversal follows" in w for w in loaded.warnings)


def test_writing_the_results(tmp_path):
    case, deviations = harmonise_tandem_speeds(example_case())
    base = base_results(case)
    results = sequence(case, base, case.settings.pacing)
    analyses = analyse_sequence(results, case.settings.gap_min, case.line)
    curve = gap_vs_pacing(case, base)
    best = min_feasible_pacing(case, base, curve)

    path = write_results(tmp_path / "out.xlsx", case, results, analyses, deviations, curve, best)
    wb = load_workbook(path)
    assert {"Summary", "Events", "Occupancy", "Gap", "PacingCurve", "MassBalance"} <= set(
        wb.sheetnames
    )
    assert wb["Events"].max_row > 10


def test_the_json_report_carries_segments_and_outcomes():
    case, deviations = harmonise_tandem_speeds(example_case())
    base = base_results(case)
    results = sequence(case, base, case.settings.pacing)
    analyses = analyse_sequence(results, case.settings.gap_min, case.line)

    report = report_to_dict(case, results, analyses, deviations)
    json.dumps(report)  # must be serialisable

    assert report["pieces"][0]["head_segments"][0]["t0_s"] == 0.0
    assert report["pieces"][0]["coiler_id"] in {"DC1", "DC2"}
    assert report["pieces"][0]["length_kinematic_m"] == pytest.approx(
        report["pieces"][0]["length_geometric_m"], rel=1e-6
    )
    assert report["gaps"][0]["ok"] is True


def test_tracking_csv_import():
    rows = [
        "piece_id,time_s,head_m,tail_m",
        "A1,0.0,0.0,-10.5",
        "A1,1.0,1.2,-9.3",
        "A2,0.0,0.0,-10.5",
    ]
    series = parse_tracking(rows)
    assert len(series) == 2
    first = next(s for s in series if s.piece_id == "A1")
    assert first.t == [0.0, 1.0]
    assert first.head == [0.0, 1.2]
    assert first.shift(10.0).t == [10.0, 11.0]


def test_tracking_without_the_tail_column():
    series = parse_tracking(["piece_id,time_s,head_m", "A1,0.0,5.0"])
    assert series[0].tail == [5.0]


def test_tracking_with_missing_columns():
    with pytest.raises(ValueError, match="missing columns"):
        parse_tracking(["piece,time", "A1,0"])
